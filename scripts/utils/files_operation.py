import logging
import os
import csv
import gzip
import time
import openpyxl
import requests
import pandas as pd
import re
from io import BytesIO, StringIO
from requests.exceptions import Timeout

# Fonction pour télécharger un fichier JSON, CSV ou Excel
def load_from_url(url, dtype=None, columns_to_keep=None, num_retries=3, delay_between_retries=5):
    logger = logging.getLogger(__name__)
    for attempt in range(num_retries):
        try:
            response = requests.get(url)
            content_type = response.headers.get('content-type')
            
            if 'json' in content_type:
                return load_json(url)
            elif 'csv' in content_type:
                return load_csv(url, dtype, columns_to_keep)
            elif re.search(r'(excel|spreadsheet|xls|xlsx)', content_type, re.IGNORECASE) or url.endswith(('.xls', '.xlsx')):
                return load_excel(url, dtype, columns_to_keep)
            else:
                logger.warning(f"Type de fichier non pris en charge à l'URL : {url}")
        except Timeout:
                handle_timeout(attempt, num_retries, delay_between_retries)
                if attempt >= num_retries - 1:
                    break
        except Exception as e:
            logger.error(f"Erreur lors du téléchargement du fichier CSV à l'URL : {url}")
            logger.error(f"Erreur : {e}")
            break
    return None 

def load_json(url):
    logger = logging.getLogger(__name__)
    response = requests.get(url)
    logger.info(f"Le fichier au format JSON a été téléchargé avec succès à l'URL : {url}")
    return response.json()

def load_csv(url, dtype=None, columns_to_keep=None):
    logger = logging.getLogger(__name__)
    response = requests.get(url)
    content = response.content
    if 'gzip' in response.headers.get('content-type'):
        content = gzip.decompress(content)
    decoded_content = content.decode('utf-8')
    delimiter = detect_delimiter(decoded_content)
    if columns_to_keep is not None:
        df = pd.read_csv(StringIO(decoded_content), delimiter=delimiter, dtype=dtype, usecols=lambda c: c in columns_to_keep, error_bad_lines=False, quoting=csv.QUOTE_MINIMAL)
    else:
        df = pd.read_csv(StringIO(decoded_content), delimiter=delimiter, dtype=dtype, error_bad_lines=False, quoting=csv.QUOTE_MINIMAL)
    logger.info(f"Le fichier CSV a été téléchargé avec succès à l'URL : {url}")
    return df

# Fonction pour détecter le délimiteur d'un fichier CSV
def detect_delimiter(text, num_lines=5, delimiters=None):
    if delimiters is None:
        delimiters = [',', ';', '\t', '|']

    counts = {delimiter: 0 for delimiter in delimiters}
    line_counts = {delimiter: 0 for delimiter in delimiters}

    for line_number, line in enumerate(StringIO(text)):
        if line_number >= num_lines:
            break
        for delimiter in delimiters:
            if delimiter in line:
                counts[delimiter] += line.count(delimiter)
                line_counts[delimiter] += 1

    # Compute average occurrences per line for each delimiter
    averages = {delimiter: counts[delimiter] / line_counts[delimiter] for delimiter in delimiters if line_counts[delimiter] > 0}

    return max(averages, key=averages.get)

def load_excel(url, dtype=None, columns_to_keep=None):
    logger = logging.getLogger(__name__)
    response = requests.get(url)
    content = BytesIO(response.content)
    workbook = openpyxl.load_workbook(content)
    sheet = workbook.active
    skiprows = detect_skiprows(sheet)
    
    skipcolumns = detect_skipcolumns(sheet)
    for _ in range(skipcolumns):
        sheet.delete_cols(1)

    temp_content = BytesIO()
    workbook.save(temp_content)
    temp_content.seek(0)

    if columns_to_keep is not None:
        df = pd.read_excel(temp_content, dtype=dtype, skiprows=skiprows, usecols=lambda c: c in columns_to_keep)
    else:
        df = pd.read_excel(temp_content, dtype=dtype, skiprows=skiprows)
    logger.info(f"Le fichier Excel a été téléchargé avec succès à l'URL : {url}")
    return df

def detect_skiprows(sheet):
    last_non_empty_col = max((cell.col_idx for row in sheet.iter_rows() for cell in row if cell.value is not None))
    first_row = 1
    for row in sheet.iter_rows(min_row=1, min_col=last_non_empty_col, max_col=last_non_empty_col):
        if row[0].value is None:
            continue
        else:
            first_row = row[0].row
            break
    return first_row - 1  # Retournez le nombre de lignes à ignorer

def detect_skipcolumns(sheet):
    last_non_empty_row = max((cell.row for row in sheet.iter_rows() for cell in row if cell.value is not None))
    first_col = 1
    for col in sheet.iter_cols(min_row=last_non_empty_row, min_col=1, max_row=last_non_empty_row):
        if col[0].value is None:
            continue
        else:
            first_col = col[0].col_idx
            break
    return first_col - 1  # Retournez le nombre de colonnes à ignorer

def handle_timeout(attempt, num_retries, delay_between_retries):
    logger = logging.getLogger(__name__)
    if attempt < num_retries - 1:
        logger.warning(f"Le téléchargement a échoué en raison d'un timeout. Tentative de réessai après {delay_between_retries} secondes...")
        time.sleep(delay_between_retries)
    else:
        logger.error(f"Le téléchargement a échoué après {num_retries} tentatives en raison d'un timeout.")

# Fonction pour télécharger et traiter plusieurs fichiers CSV
def download_and_process_data(urls_list, dtype=None):
    data_frames = []
    for url in urls_list:
        df = load_from_url(url, dtype=dtype)
        if df is not None:
            data_frames.append(df)
    return data_frames

# Téléchargement des fichiers Excel 
def load_from_path(file_path, dtype=None):
    return pd.read_excel(file_path, dtype=dtype)

def save_csv(df, file_folder, file_name, sep=","):
    logger = logging.getLogger(__name__)
    # Vérifie si le répertoire existe, le crée si nécessaire
    if not os.path.exists(file_folder):
        os.makedirs(file_folder)

    df.to_csv(file_folder / file_name, index=False, sep=sep)
    logger.info(f"Le fichier {file_name} a été enregistré dans le répertoire {file_folder}")